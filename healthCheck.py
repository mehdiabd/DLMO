__author__ = "S. Mehdi Abdollahi"

from pymongo.errors import BulkWriteError
from datetime import datetime, timedelta
from ipaddress import ip_address
from pymongo import MongoClient
from openpyxl import Workbook

# Main
if __name__ == '__main__':

    wb = Workbook()
    ws = wb.active
    ws.title = "DLMO Health Check"
    ws.sheet_properties.tabColor = "000000"

    def get_short_date(days_interval=0, sdate=None):
        sdate = datetime.strptime(str(sdate), '%y%m%d') if sdate else datetime.today()
        return int((sdate + timedelta(days=days_interval)).strftime('%y%m%d'))
    gsd = get_short_date(-1)

# DB configuration
    dcf_db = MongoClient("mongodb://172.30.96.230:27017", connect=False)['dcf']
    adf_db = MongoClient("mongodb://172.30.96.235:27017,172.30.96.236:27017,"
                         "172.30.96.237:27017/?replicaSet=ADFrs&readPreference=secondaryPreferred", connect=False)[
        'adf']
    dlm_db = MongoClient("mongodb://172.30.96.132:27017", connect=False)['dlm']
    bulk_op = dlm_db.get_collection('res_cl').initialize_unordered_bulk_op()

# IPs
    pilot_ips = [
        "172.31.18.4",
        "172.31.18.5",
        "172.31.18.6",
        "172.31.18.7",
        "172.31.18.8",
        "172.31.18.9",
        "172.31.18.11",
        "172.31.18.12",
        "172.31.18.13",
        "172.31.18.14",
        "172.31.18.15",
        "172.31.18.16",
        "172.31.18.17",
        "172.31.18.18",
        "172.31.18.19",
        "172.31.18.20",
        "172.31.18.21",
        "172.31.18.22",
        "172.31.18.23",
        "172.31.18.24",
        "172.31.18.25",
        "172.31.18.26",
        "172.31.18.28",
        "172.31.114.4",
        "172.31.114.5",
        "172.31.114.6",
        "172.31.114.7",
        "172.31.114.8",
        "172.31.114.9",
        "172.31.114.11",
        "172.31.114.12",
        "172.31.114.13",
        "172.31.114.15",
        "172.31.114.16",
        "172.31.114.19",
        "172.31.114.23",
        "172.31.114.24",
        "172.31.114.25",
        "172.31.114.27",
        "172.31.114.28",
        "172.31.114.29",
        "172.31.114.30",
        "172.31.114.31",
        "172.31.114.32",
        "172.31.114.33",
        "172.31.114.34",
        "172.31.114.35",
        "172.31.114.36",
        "172.31.114.37",
        "172.31.114.39",
        "172.31.131.132",
        "172.31.131.134",
        "172.31.131.135",
        "172.31.131.136",
        "172.31.131.138",
        "172.31.131.139",
        "172.31.131.140",
        "172.31.131.141",
        "172.31.64.4",
        "172.31.64.5",
        "172.31.64.6",
        "172.31.64.7",
        "172.31.64.8",
        "172.31.64.9",
        "172.31.64.10",
        "172.31.64.11",
        "172.31.64.12",
        "172.31.64.13",
        "172.31.64.14",
        "172.31.64.15",
        "172.31.64.16",
        "172.31.64.17",
        "172.31.64.18",
        "172.31.64.19",
        "172.31.64.20",
        "172.31.64.21",
        "172.31.64.22",
        "172.31.64.23",
        "172.31.48.50",
        "172.31.48.51",
        "172.31.48.4",
        "172.31.48.5",
        "172.31.48.6",
        "172.31.48.47",
        "172.31.48.48",
        "172.31.48.49",
        "172.31.48.31",
        "172.31.48.32",
        "172.31.48.33",
        "172.31.48.27",
        "172.31.48.8",
        "172.31.48.23",
        "172.31.48.22",
        "172.31.48.7",
        "172.31.48.12",
        "172.31.48.13",
        "172.31.48.24",
        "172.31.84.133",
        "172.31.84.134",
        "172.31.84.142",
        "172.31.84.147",
        "172.31.84.148",
        "172.31.84.149",
        "172.31.84.150",
        "172.31.84.151",
        "172.31.84.152",
        "172.31.84.153",
        "172.31.84.154",
        "172.31.84.155",
        "172.31.84.156",
        "172.31.84.135",
        "172.31.84.144",
        "172.31.84.138",
        "172.31.84.139",
        "172.31.84.137",
        "172.31.84.145",
        "172.31.84.143",
        "172.31.84.136",
        "172.31.34.4",
        "172.31.34.5",
        "172.31.34.6",
        "172.31.34.29",
        "172.31.34.30",
        "172.31.34.31",
        "172.31.34.42",
        "172.31.34.35",
        "172.31.34.36",
        "172.31.34.37",
        "172.31.34.8",
        "172.31.34.51",
        "172.31.34.9",
        "172.31.34.49",
        "172.31.34.47",
        "172.31.34.7",
        "172.31.34.48",
        "172.31.34.50",
        "172.31.34.53"
    ]
    # Chitgar IP
    # pilot_ips = ["10.41.16.17"]

    print("Checking the System Health ...")

    ws.append(["IP", "DCF p/n", "DCF Bucket", "DCF Singular", "ADF Line Quality", "ADF Weighted Quality",
               "DLM Selection Prequalify", "DLM Initial Profile Selection", "DLM NE Profile Sync",
               "DLM Profile Configuration Function", "SP InProgress", "SP Stabilize", "SP Optimize", "SP TOTAL",
               "IPS Waiting", "IPS Initial", "IPS Operation", "IPS OptimizeImpossible", "IPS Unstable", "IPS TOTAL",
               "PCF Success", "PCF Roll Back", "PCF Line Down", "PCF Retry Limit Reach", "PCF C Profile N/A",
               "PCF TOTAL"])

    counter = 2
    db_row_counter = 0

    def cursor(db, clt, par1, par2):
        return db[clt].find_one({par1: par2, "sdate": gsd})

    def ui():
        while True:
            option = input("Type of report (db/xl) or 'ext' to exit app: ")

            if option != 'db' and option != 'xl' and option != 'ext':
                print('Invalid type.')
                #continue

            elif option == 'db':
                if db_row_counter != 0:
                    try:
                        bulk_op.execute()
                        print('Everything OK')
                    except BulkWriteError as bwe:
                        print(bwe.details)
                        raise
                else:
                    print("Nothing Upserted.")
                #continue

            elif option == 'xl':
                wb.save("/home/mehdi/PycharmProjects/DLMO/healthCheck reports/v9/rpt-" + str(gsd) + ".xlsx")
                print("See results in: /home/mehdi/PycharmProjects/DLMO/healthCheck reports . ")
                #continue

            elif option == 'ext':
                print('Bye.')
                break


    for ip in pilot_ips:

        mod_dict = {}
        ipsDict = {}
        spDict = {}
        pcfDict = {}
        ws.append([ip])
        collection = "ne" + str(int(ip_address(ip)))

        dcf_pn_res = cursor(dcf_db, collection, "ptype", 'p/n')
        if not dcf_pn_res:
            ws.cell(row=counter, column=2, value='Null')
            mod_dict["DCF p/n"] = "Null"
        else:
            ws.cell(row=counter, column=2, value='OK')
            mod_dict["DCF p/n"] = "OK"

        dcf_bucket_res = cursor(dcf_db, collection, "ptype", 'b/i')
        if not dcf_bucket_res:
            for i in range(3, 7):
                ws.cell(row=counter, column=i, value='Null')
            mod_dict.update({"DCF Bucket": "Null", "DCF Singular": "Null", "ADF Line Quality": "Null",
                             "ADF Weighted Quality": "Null"})
        else:
            ws.cell(row=counter, column=3, value='OK')
            mod_dict["DCF Bucket"] = "OK"

            dcf_singular_res = cursor(dcf_db, collection, "ptype", 's/i')
            if not dcf_singular_res:
                for i in range(4, 7):
                    ws.cell(row=counter, column=i, value='Null')
                mod_dict.update({"DCF Singular": "Null", "ADF Line Quality": "Null", "ADF Weighted Quality": "Null"})
            else:
                ws.cell(row=counter, column=4, value='OK')
                mod_dict["DCF Singular"] = "OK"

            adf_res_lq = cursor(adf_db, 'line_quality', 'ip_adr', ip)
            if not adf_res_lq:
                for i in range(5, 7):
                    ws.cell(row=counter, column=i, value='Null')
                mod_dict.update({"ADF Line Quality": "Null", "ADF Weighted Quality": "Null"})
            else:
                ws.cell(row=counter, column=5, value='OK')
                mod_dict["ADF Line Quality"] = "OK"

                adf_res_wq = cursor(adf_db, 'weighted_quality', 'ip_adr', ip)
                if not adf_res_wq:
                    ws.cell(row=counter, column=6, value='Null')
                    mod_dict["ADF Weighted Quality"] = "Null"
                else:
                    ws.cell(row=counter, column=6, value='OK')
                    mod_dict["ADF Weighted Quality"] = "OK"

        adf_wq_2 = adf_db['weighted_quality'].find_one({'ip_adr': ip, "sdate": get_short_date(-2)})
        if not adf_wq_2:
            for i in range(7, 11):
                ws.cell(row=counter, column=i, value='Null')
            mod_dict.update({"DLM Selection Prequalify": "Null", "DLM Initial Profile Selection": "Null",
                             "DLM NE Profile Sync": "Null", "DLM Profile Configuration Function": "Null"})
            pass
        else:
            dlm_res_sp = cursor(dlm_db, 'selection_prequalify', 'ip_adr', ip)
            if not dlm_res_sp:
                for i in range(7, 11):
                    ws.cell(row=counter, column=i, value='Null')
                mod_dict.update({"DLM Selection Prequalify": "Null", "DLM Initial Profile Selection": "Null",
                                 "DLM NE Profile Sync": "Null", "DLM Profile Configuration Function": "Null"})
            else:
                ws.cell(row=counter, column=7, value='OK')
                mod_dict["DLM Selection Prequalify"] = "OK"

                dlm_sp_res = dlm_db['selection_prequalify'].aggregate(
                    [
                        {"$match": {"sdate": gsd, "ip_adr": ip}},
                        {"$project": {"state": 1}},
                        {
                            "$group":
                                {
                                    "_id": "$state"
                                    , "count": {"$sum": 1}
                                }
                        }
                    ]
                )

                dlm_sp_res = list(dlm_sp_res)
                for id in dlm_sp_res:
                    if id['_id'] == "InProgress":
                        spDict["InProgress"] = id['count']
                        mod_dict["SP InProgress No"] = id['count']
                    elif id['_id'] == "Stabilize":
                        spDict['Stabilize'] = id['count']
                        mod_dict["SP Stabilize No"] = id['count']
                    else:
                        spDict['Optimize'] = id['count']
                        mod_dict["SP Optimize No"] = id['count']

                dlm_res_ips = cursor(dlm_db, 'initial_profile_selection', 'ip_adr', ip)
                if not dlm_res_ips:
                    for i in range(8, 11):
                        ws.cell(row=counter, column=i, value='Null')
                    mod_dict.update({"DLM Initial Profile Selection": "Null", "DLM NE Profile Sync": "Null",
                                     "DLM Profile Configuration Function": "Null"})
                else:
                    ws.cell(row=counter, column=8, value='OK')
                    mod_dict["DLM Initial Profile Selection"] = 'OK'

                    dlm_ips_res = dlm_db['initial_profile_selection'].aggregate(
                        [
                            {"$match": {"sdate": gsd, "ip_adr": ip}},
                            {"$project": {"state": 1}},
                            {
                                "$group":
                                    {
                                        "_id": "$state"
                                        , "count": {"$sum": 1}
                                    }
                            }
                        ]
                    )

                    dlm_ips_res = list(dlm_ips_res)
                    for i in dlm_ips_res:
                        if i['_id'] == "Waiting":
                            ipsDict["Waiting"] = i['count']
                            mod_dict["IPS Waiting No"] = i['count']
                        elif i['_id'] == "Initial":
                            ipsDict["Initial"] = i['count']
                            mod_dict["IPS Initial No"] = i['count']
                        elif i['_id'] == "Operation":
                            ipsDict["Operation"] = i['count']
                            mod_dict["IPS Operation No"] = i['count']
                        elif i['_id'] == "OptimizeImpossible":
                            ipsDict["OptimizeImpossible"] = i['count']
                            mod_dict["IPS OptimizeImpossible No"] = i['count']
                        else:
                            ipsDict['Unstable'] = i['count']
                            mod_dict['IPS Unstable No'] = i['count']

                    dlm_res_nps = cursor(dlm_db, 'ne_profile_sync', 'ip_adr', ip)
                    if not dlm_res_nps:
                        ws.cell(row=counter, column=9, value='Null')
                        mod_dict['DLM NE Profile Sync'] = 'Null'
                    else:
                        ws.cell(row=counter, column=9, value='OK')
                        mod_dict['DLM NE Profile Sync'] = 'OK'

                    dlm_res_pcf = cursor(dlm_db, 'profile_configuration_function', 'ip_adr', ip)
                    if not dlm_res_pcf:
                        ws.cell(row=counter, column=10, value='Null')
                        mod_dict['DLM Profile Configuration Function'] = 'Null'
                    else:
                        ws.cell(row=counter, column=10, value='OK')
                        mod_dict['DLM Profile Configuration Function'] = 'OK'

                        pcf_res_se = dlm_db['profile_configuration_function'].aggregate(
                            [
                                {"$match": {"sdate": gsd, "ip_adr": ip}},
                                {"$project": {"state": 1, "error": 1}},
                                {
                                    "$group":
                                        {
                                            "_id": {"stat": "$state", "erro": "$error"}
                                            , "count": {"$sum": 1}
                                        }
                                }
                            ]
                        )

                        pcf_res_se = list(pcf_res_se)
                        for id in pcf_res_se:
                            if id['_id']['stat'] == "Success":
                                pcfDict["Success"] = id['count']
                                mod_dict["PCF Success No"] = id['count']
                            else:
                                pcfDict[id['_id']['erro']] = id['count']
                                mod_dict[id['_id']['erro']] = id['count']

# DLM SP
        g = int(spDict.get('InProgress', 0))
        h = int(spDict.get('Stabilize', 0))
        i = int(spDict.get('Optimize', 0))
# DLM IPS
        j = int(ipsDict.get('Waiting', 0))
        k = int(ipsDict.get('Initial', 0))
        l = int(ipsDict.get('Operation', 0))
        m = int(ipsDict.get('OptimizeImpossible', 0))
        a = int(ipsDict.get('Unstable', 0))
# DLM PCF
        b = int(pcfDict.get('Success', 0))
        c = int(pcfDict.get('rolled back', 0))
        d = int(pcfDict.get('line is down', 0))
        e = int(pcfDict.get('retry limit reached!', 0))
        f = int(pcfDict.get('current profile is not available', 0))

# Worksheet insertions
        ws.cell(row=counter, column=11, value=g)
        ws.cell(row=counter, column=12, value=h)
        ws.cell(row=counter, column=13, value=i)
        ws[f"N{counter}"] = f'= SUM(K{counter}:M{counter})'
        ws.cell(row=counter, column=15, value=j)
        ws.cell(row=counter, column=16, value=k)
        ws.cell(row=counter, column=17, value=l)
        ws.cell(row=counter, column=18, value=m)
        ws.cell(row=counter, column=19, value=a)
        ws[f"T{counter}"] = f'= SUM(O{counter}:S{counter})'
        ws.cell(row=counter, column=21, value=b)
        ws.cell(row=counter, column=22, value=c)
        ws.cell(row=counter, column=23, value=d)
        ws.cell(row=counter, column=24, value=e)
        ws.cell(row=counter, column=25, value=f)
        ws[f"Z{counter}"] = f'= SUM(U{counter}:Y{counter})'

        print(mod_dict)
        bulk_op.find({"ip": ip, "sdate": gsd}).upsert().update_one({"$set": mod_dict})
        db_row_counter += 1

        counter += 1

    ui()
